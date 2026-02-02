from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.core.exceptions import ValidationError
from .forms import MataKuliahForm
from .models import MataKuliah
import json
import csv
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

@login_required(login_url='/admin/login/')
@never_cache
def input_matakuliah(request):
    pesan = None
    if request.method == 'POST':
        form = MataKuliahForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    instance = form.save()
                messages.success(request, f'Data mata kuliah "{instance.nama_mk}" berhasil ditambahkan!', extra_tags='success')
                return redirect('input_matakuliah')
            except ValidationError as e:
                messages.error(request, f'Gagal menambahkan data: {str(e)}', extra_tags='danger')
            except Exception as e:
                messages.error(request, f'Terjadi kesalahan: {str(e)}', extra_tags='danger')
        else:
            # Form tidak valid
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            error_text = " | ".join(error_messages) if error_messages else "Terjadi kesalahan pada form"
            messages.error(request, f'Gagal menambahkan data: {error_text}', extra_tags='danger')
    else:
        form = MataKuliahForm()

    matakuliah = MataKuliah.objects.all().order_by('id')
    
    # Search
    search_query = request.GET.get('q', '')
    
    if search_query:
        from django.db.models import Q
        matakuliah = matakuliah.filter(Q(nama_mk__icontains=search_query) | Q(kode_mk__icontains=search_query))

    return render(request, 'matakuliah/input.html', {
        'form': form,
        'matakuliah': matakuliah,
        'pesan': pesan,
        'search_query': search_query,
    })


@login_required(login_url='/admin/login/')
@require_POST
@never_cache
def matakuliah_delete(request, pk):
    try:
        with transaction.atomic():
            mk = MataKuliah.objects.get(pk=pk)
            nama = mk.nama_mk
            mk.delete()
        messages.success(request, f'Data mata kuliah "{nama}" berhasil dihapus!', extra_tags='success')
        return JsonResponse({'success': True})
    except MataKuliah.DoesNotExist:
        messages.error(request, 'Data tidak ditemukan', extra_tags='danger')
        return JsonResponse({'success': False, 'error': 'Data tidak ditemukan'})
    except Exception as e:
        messages.error(request, f'Gagal menghapus data: {str(e)}', extra_tags='danger')
        return JsonResponse({'success': False, 'error': str(e)})


@login_required(login_url='/admin/login/')
@require_POST
@never_cache
def matakuliah_update(request, pk):
    try:
        data = json.loads(request.body)
        matakuliah = get_object_or_404(MataKuliah, pk=pk)
        
        # Update fields
        matakuliah.nama_mk = data.get('nama_mk', matakuliah.nama_mk)
        matakuliah.kode_mk = data.get('kode_mk', matakuliah.kode_mk)
        matakuliah.sks = data.get('sks', matakuliah.sks)
        matakuliah.semester = data.get('semester', matakuliah.semester)
        matakuliah.dosen_mk_id = data.get('dosen_mk', matakuliah.dosen_mk_id)
        
        # Validate and save
        matakuliah.full_clean()
        matakuliah.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Data berhasil diupdate'
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON format'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required(login_url='/admin/login/')
@require_http_methods(["GET"])
def export_matakuliah_csv(request):
    """Export matakuliah data to CSV"""
    matakuliah = MataKuliah.objects.all().order_by('id')
    
    search_query = request.GET.get('q', '')
    if search_query:
        from django.db.models import Q
        matakuliah = matakuliah.filter(Q(nama_mk__icontains=search_query) | Q(kode_mk__icontains=search_query))
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="matakuliah_{timestamp}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['No', 'Nama MK', 'Kode MK', 'SKS', 'Semester', 'Dosen', 'Jumlah Mahasiswa'])
    
    for index, mk in enumerate(matakuliah, 1):
        writer.writerow([
            index,
            mk.nama_mk,
            mk.kode_mk,
            mk.sks,
            mk.semester,
            mk.dosen_mk.nama if mk.dosen_mk else '',
            mk.mhs_mk.all().count()
        ])
    
    return response


@login_required(login_url='/admin/login/')
@require_http_methods(["GET"])
def export_matakuliah_excel(request):
    """Export matakuliah data to Excel"""
    if not HAS_OPENPYXL:
        return HttpResponse("openpyxl library is not installed", status=400)
    
    matakuliah = MataKuliah.objects.all().order_by('id')
    
    search_query = request.GET.get('q', '')
    if search_query:
        from django.db.models import Q
        matakuliah = matakuliah.filter(Q(nama_mk__icontains=search_query) | Q(kode_mk__icontains=search_query))
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Mata Kuliah"
    
    # Header styling
    header_fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
    header_font = Font(bold=True, color="333333")
    
    headers = ['No', 'Nama MK', 'Kode MK', 'SKS', 'Semester', 'Dosen', 'Jumlah Mahasiswa']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, mk in enumerate(matakuliah, 2):
        ws.cell(row=row_num, column=1).value = row_num - 1
        ws.cell(row=row_num, column=2).value = mk.nama_mk
        ws.cell(row=row_num, column=3).value = mk.kode_mk
        ws.cell(row=row_num, column=4).value = mk.sks
        ws.cell(row=row_num, column=5).value = mk.semester
        ws.cell(row=row_num, column=6).value = mk.dosen_mk.nama if mk.dosen_mk else ''
        ws.cell(row=row_num, column=7).value = mk.mhs_mk.all().count()
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 18
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="matakuliah_{timestamp}.xlsx"'
    wb.save(response)
    
    return response