from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.core.exceptions import ValidationError
from .forms import DosenForm
from .models import Dosen
import json
import csv
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

@login_required(login_url='/admin/login/')
@never_cache
def input_dosen(request):
    pesan = None
    if request.method == 'POST':
        form = DosenForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    instance = form.save()
                messages.success(request, f'Data dosen "{instance.nama}" berhasil ditambahkan!', extra_tags='success')
                return redirect('input_dosen')
            except ValidationError as e:
                messages.error(request, f'Gagal menambahkan data: {str(e)}', extra_tags='danger')
            except Exception as e:
                messages.error(request, f'Terjadi kesalahan: {str(e)}', extra_tags='danger')
        else:
            # Form tidak valid
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            error_text = " | ".join(error_messages) if error_messages else "Terjadi kesalahan pada form"
            messages.error(request, f'Gagal menambahkan data: {error_text}', extra_tags='danger')
    else:
        form = DosenForm()

    dosen = Dosen.objects.all().order_by('id')
    
    # Search
    search_query = request.GET.get('q', '')
    
    if search_query:
        from django.db.models import Q
        dosen = dosen.filter(Q(nama__icontains=search_query) | Q(nidn__icontains=search_query))

    return render(request, 'dosen/input.html', {
        'form': form,
        'dosen': dosen,
        'pesan': pesan,
        'search_query': search_query,
    })


@login_required(login_url='/admin/login/')
@require_POST
@never_cache
def dosen_delete(request, pk):
    try:
        with transaction.atomic():
            d = Dosen.objects.get(pk=pk)
            nama = d.nama
            d.delete()
        messages.success(request, f'Data dosen "{nama}" berhasil dihapus!', extra_tags='success')
        return JsonResponse({'success': True})
    except Dosen.DoesNotExist:
        messages.error(request, 'Data tidak ditemukan', extra_tags='danger')
        return JsonResponse({'success': False, 'error': 'Data tidak ditemukan'})
    except Exception as e:
        messages.error(request, f'Gagal menghapus data: {str(e)}', extra_tags='danger')
        return JsonResponse({'success': False, 'error': str(e)})


@login_required(login_url='/admin/login/')
@require_POST
@never_cache
def dosen_update(request, pk):
    try:
        data = json.loads(request.body)
        dosen = get_object_or_404(Dosen, pk=pk)
        
        # Update fields
        dosen.nama = data.get('nama', dosen.nama)
        dosen.nidn = data.get('nidn', dosen.nidn)
        dosen.email = data.get('email', dosen.email)
        dosen.no_hp = data.get('no_hp', dosen.no_hp)
        dosen.alamat = data.get('alamat', dosen.alamat)
        dosen.homebase = data.get('homebase', dosen.homebase)
        
        # Validate and save
        dosen.full_clean()
        dosen.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Data berhasil diupdate'
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON format'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required(login_url='/admin/login/')
@require_http_methods(["GET"])
def export_dosen_csv(request):
    """Export dosen data to CSV"""
    dosen = Dosen.objects.all().order_by('id')
    
    search_query = request.GET.get('q', '')
    if search_query:
        from django.db.models import Q
        dosen = dosen.filter(Q(nama__icontains=search_query) | Q(nidn__icontains=search_query))
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="dosen_{timestamp}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['No', 'Nama', 'NIDN', 'Email', 'No. HP', 'Homebase', 'Alamat'])
    
    for index, d in enumerate(dosen, 1):
        writer.writerow([
            index,
            d.nama,
            d.nidn,
            d.email,
            d.no_hp or '',
            d.homebase or '',
            d.alamat or ''
        ])
    
    return response


@login_required(login_url='/admin/login/')
@require_http_methods(["GET"])
def export_dosen_excel(request):
    """Export dosen data to Excel"""
    if not HAS_OPENPYXL:
        return HttpResponse("openpyxl library is not installed", status=400)
    
    dosen = Dosen.objects.all().order_by('id')
    
    search_query = request.GET.get('q', '')
    if search_query:
        from django.db.models import Q
        dosen = dosen.filter(Q(nama__icontains=search_query) | Q(nidn__icontains=search_query))
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Dosen"
    
    # Header styling
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ['No', 'Nama', 'NIDN', 'Email', 'No. HP', 'Homebase', 'Alamat']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, d in enumerate(dosen, 2):
        ws.cell(row=row_num, column=1).value = row_num - 1
        ws.cell(row=row_num, column=2).value = d.nama
        ws.cell(row=row_num, column=3).value = d.nidn
        ws.cell(row=row_num, column=4).value = d.email
        ws.cell(row=row_num, column=5).value = d.no_hp or ''
        ws.cell(row=row_num, column=6).value = d.homebase or ''
        ws.cell(row=row_num, column=7).value = d.alamat or ''
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="dosen_{timestamp}.xlsx"'
    wb.save(response)
    
    return response