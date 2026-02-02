from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Count
from django.core.exceptions import ValidationError
from .forms import MahasiswaForm
from .models import Mahasiswa
from dosen.models import Dosen
from matakuliah.models import MataKuliah
import json
import csv
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

@login_required(login_url='/admin/login/')
@never_cache
def tampilkan_semua_data(request):
    mahasiswa = Mahasiswa.objects.all().order_by('id')
    dosen = Dosen.objects.all().order_by('id')
    matakuliah = MataKuliah.objects.all().order_by('id')
    
    return render(request, 'tampilkan_semua_data.html', {
        'mahasiswa': mahasiswa,
        'dosen': dosen,
        'matakuliah': matakuliah,
    })

@login_required(login_url='/admin/login/')
@never_cache
def input_mahasiswa(request):
    pesan = None
    if request.method == 'POST':
        form = MahasiswaForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Call full_clean() untuk memastikan model-level validation
                    instance = form.save(commit=False)
                    instance.full_clean()  # Model validation
                    instance.save()  # Save ke database
                messages.success(request, f'Data mahasiswa "{instance.nama}" berhasil ditambahkan!', extra_tags='success')
                return redirect('input_mahasiswa')  # Redirect after POST
            except ValidationError as e:
                messages.error(request, f'Gagal menambahkan data: {str(e)}', extra_tags='danger')
            except Exception as e:
                messages.error(request, f'Terjadi kesalahan: {str(e)}', extra_tags='danger')
        else:
            # Form tidak valid
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            error_text = " | ".join(error_messages) if error_messages else "Terjadi kesalahan pada form"
            messages.error(request, f'Gagal menambahkan data: {error_text}', extra_tags='danger')
    else:
        form = MahasiswaForm()

    mahasiswa = Mahasiswa.objects.all().order_by('id')
    
    # Search and Filter
    search_query = request.GET.get('q', '')
    jurusan_filter = request.GET.get('jurusan', '')

    if search_query:
        from django.db.models import Q
        mahasiswa = mahasiswa.filter(Q(nama__icontains=search_query) | Q(npm__icontains=search_query))
    
    if jurusan_filter:
        mahasiswa = mahasiswa.filter(jurusan=jurusan_filter)

    return render(request, 'mahasiswa/input.html', {
        'form': form,
        'mahasiswa': mahasiswa,
        'pesan': pesan,
        'search_query': search_query,
        'jurusan_filter': jurusan_filter,
    })

@login_required(login_url='/admin/login/')
@require_POST
@never_cache
def mahasiswa_delete(request, pk):
    try:
        with transaction.atomic():
            m = Mahasiswa.objects.get(pk=pk)
            nama = m.nama
            m.delete()
        messages.success(request, f'Data mahasiswa "{nama}" berhasil dihapus!', extra_tags='success')
        return JsonResponse({'success': True})
    except Mahasiswa.DoesNotExist:
        messages.error(request, 'Data tidak ditemukan', extra_tags='danger')
        return JsonResponse({'success': False, 'error': 'Data tidak ditemukan'})
    except Exception as e:
        messages.error(request, f'Gagal menghapus data: {str(e)}', extra_tags='danger')
        return JsonResponse({'success': False, 'error': str(e)})

@login_required(login_url='/admin/login/')
@require_POST
@never_cache
def mahasiswa_update(request, pk):
    try:
        data = json.loads(request.body)
        mahasiswa = get_object_or_404(Mahasiswa, pk=pk)
        
        # Update fields
        mahasiswa.nama = data.get('nama', mahasiswa.nama)
        mahasiswa.npm = data.get('npm', mahasiswa.npm)
        mahasiswa.email = data.get('email', mahasiswa.email)
        mahasiswa.no_hp = data.get('no_hp', mahasiswa.no_hp)
        mahasiswa.jurusan = data.get('jurusan', mahasiswa.jurusan)
        mahasiswa.alamat = data.get('alamat', mahasiswa.alamat)
        
        # Validate and save
        mahasiswa.full_clean()
        mahasiswa.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Data berhasil diupdate'
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON format'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required(login_url='/admin/login/')
@require_http_methods(["GET"])
def export_mahasiswa_csv(request):
    """Export mahasiswa data to CSV"""
    # Get filtered data if search/filter exists
    mahasiswa = Mahasiswa.objects.all().order_by('id')
    
    search_query = request.GET.get('q', '')
    jurusan_filter = request.GET.get('jurusan', '')

    if search_query:
        from django.db.models import Q
        mahasiswa = mahasiswa.filter(Q(nama__icontains=search_query) | Q(npm__icontains=search_query))
    
    if jurusan_filter:
        mahasiswa = mahasiswa.filter(jurusan=jurusan_filter)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="mahasiswa_{timestamp}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['No', 'Nama', 'NPM', 'Email', 'No. HP', 'Jurusan', 'Alamat'])
    
    for index, m in enumerate(mahasiswa, 1):
        writer.writerow([
            index,
            m.nama,
            m.npm,
            m.email,
            m.no_hp or '',
            m.jurusan or '',
            m.alamat or ''
        ])
    
    return response


@login_required(login_url='/admin/login/')
@require_http_methods(["GET"])
def export_mahasiswa_excel(request):
    """Export mahasiswa data to Excel"""
    if not HAS_OPENPYXL:
        return HttpResponse("openpyxl library is not installed", status=400)
    
    # Get filtered data if search/filter exists
    mahasiswa = Mahasiswa.objects.all().order_by('id')
    
    search_query = request.GET.get('q', '')
    jurusan_filter = request.GET.get('jurusan', '')

    if search_query:
        from django.db.models import Q
        mahasiswa = mahasiswa.filter(Q(nama__icontains=search_query) | Q(npm__icontains=search_query))
    
    if jurusan_filter:
        mahasiswa = mahasiswa.filter(jurusan=jurusan_filter)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Mahasiswa"
    
    # Header styling
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ['No', 'Nama', 'NPM', 'Email', 'No. HP', 'Jurusan', 'Alamat']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, m in enumerate(mahasiswa, 2):
        ws.cell(row=row_num, column=1).value = row_num - 1
        ws.cell(row=row_num, column=2).value = m.nama
        ws.cell(row=row_num, column=3).value = m.npm
        ws.cell(row=row_num, column=4).value = m.email
        ws.cell(row=row_num, column=5).value = m.no_hp or ''
        ws.cell(row=row_num, column=6).value = m.jurusan or ''
        ws.cell(row=row_num, column=7).value = m.alamat or ''
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="mahasiswa_{timestamp}.xlsx"'
    wb.save(response)
    
    return response


@login_required(login_url='/admin/login/')
@require_http_methods(["GET"])
def export_all_data_csv(request):
    """Export all data (mahasiswa, dosen, matakuliah) to CSV files"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Create a combined CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="semua_data_{timestamp}.csv"'
    
    writer = csv.writer(response)
    
    # Write Mahasiswa data
    writer.writerow(['=== DATA MAHASISWA ==='])
    writer.writerow(['No', 'Nama', 'NPM', 'Email', 'No. HP', 'Jurusan', 'Alamat'])
    mahasiswa = Mahasiswa.objects.all().order_by('id')
    for index, m in enumerate(mahasiswa, 1):
        writer.writerow([index, m.nama, m.npm, m.email, m.no_hp or '', m.jurusan or '', m.alamat or ''])
    
    writer.writerow([])
    writer.writerow(['=== DATA DOSEN ==='])
    writer.writerow(['No', 'Nama', 'NIDN', 'Email', 'No. HP', 'Homebase', 'Alamat'])
    dosen = Dosen.objects.all().order_by('id')
    for index, d in enumerate(dosen, 1):
        writer.writerow([index, d.nama, d.nidn, d.email, d.no_hp or '', d.homebase or '', d.alamat or ''])
    
    writer.writerow([])
    writer.writerow(['=== DATA MATA KULIAH ==='])
    writer.writerow(['No', 'Nama MK', 'Kode MK', 'SKS', 'Semester', 'Dosen', 'Jumlah Mahasiswa'])
    matakuliah = MataKuliah.objects.all().order_by('id')
    for index, mk in enumerate(matakuliah, 1):
        writer.writerow([index, mk.nama_mk, mk.kode_mk, mk.sks, mk.semester, 
                        mk.dosen_mk.nama if mk.dosen_mk else '', mk.mhs_mk.all().count()])
    
    return response


@login_required(login_url='/admin/login/')
@require_http_methods(["GET"])
def export_all_data_excel(request):
    """Export all data (mahasiswa, dosen, matakuliah) to Excel"""
    if not HAS_OPENPYXL:
        return HttpResponse("openpyxl library is not installed", status=400)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Create workbook with multiple sheets
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Mahasiswa sheet
    ws_mhs = wb.create_sheet("Mahasiswa")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ['No', 'Nama', 'NPM', 'Email', 'No. HP', 'Jurusan', 'Alamat']
    for col_num, header in enumerate(headers, 1):
        cell = ws_mhs.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    mahasiswa = Mahasiswa.objects.all().order_by('id')
    for row_num, m in enumerate(mahasiswa, 2):
        ws_mhs.cell(row=row_num, column=1).value = row_num - 1
        ws_mhs.cell(row=row_num, column=2).value = m.nama
        ws_mhs.cell(row=row_num, column=3).value = m.npm
        ws_mhs.cell(row=row_num, column=4).value = m.email
        ws_mhs.cell(row=row_num, column=5).value = m.no_hp or ''
        ws_mhs.cell(row=row_num, column=6).value = m.jurusan or ''
        ws_mhs.cell(row=row_num, column=7).value = m.alamat or ''
    
    ws_mhs.column_dimensions['A'].width = 5
    ws_mhs.column_dimensions['B'].width = 25
    ws_mhs.column_dimensions['C'].width = 15
    ws_mhs.column_dimensions['D'].width = 25
    ws_mhs.column_dimensions['E'].width = 15
    ws_mhs.column_dimensions['F'].width = 20
    ws_mhs.column_dimensions['G'].width = 25
    
    # Dosen sheet
    ws_dos = wb.create_sheet("Dosen")
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    
    headers = ['No', 'Nama', 'NIDN', 'Email', 'No. HP', 'Homebase', 'Alamat']
    for col_num, header in enumerate(headers, 1):
        cell = ws_dos.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    dosen = Dosen.objects.all().order_by('id')
    for row_num, d in enumerate(dosen, 2):
        ws_dos.cell(row=row_num, column=1).value = row_num - 1
        ws_dos.cell(row=row_num, column=2).value = d.nama
        ws_dos.cell(row=row_num, column=3).value = d.nidn
        ws_dos.cell(row=row_num, column=4).value = d.email
        ws_dos.cell(row=row_num, column=5).value = d.no_hp or ''
        ws_dos.cell(row=row_num, column=6).value = d.homebase or ''
        ws_dos.cell(row=row_num, column=7).value = d.alamat or ''
    
    ws_dos.column_dimensions['A'].width = 5
    ws_dos.column_dimensions['B'].width = 25
    ws_dos.column_dimensions['C'].width = 15
    ws_dos.column_dimensions['D'].width = 25
    ws_dos.column_dimensions['E'].width = 15
    ws_dos.column_dimensions['F'].width = 20
    ws_dos.column_dimensions['G'].width = 25
    
    # Mata Kuliah sheet
    ws_mk = wb.create_sheet("Mata Kuliah")
    header_fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
    
    headers = ['No', 'Nama MK', 'Kode MK', 'SKS', 'Semester', 'Dosen', 'Jumlah Mahasiswa']
    for col_num, header in enumerate(headers, 1):
        cell = ws_mk.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    matakuliah = MataKuliah.objects.all().order_by('id')
    for row_num, mk in enumerate(matakuliah, 2):
        ws_mk.cell(row=row_num, column=1).value = row_num - 1
        ws_mk.cell(row=row_num, column=2).value = mk.nama_mk
        ws_mk.cell(row=row_num, column=3).value = mk.kode_mk
        ws_mk.cell(row=row_num, column=4).value = mk.sks
        ws_mk.cell(row=row_num, column=5).value = mk.semester
        ws_mk.cell(row=row_num, column=6).value = mk.dosen_mk.nama if mk.dosen_mk else ''
        ws_mk.cell(row=row_num, column=7).value = mk.mhs_mk.all().count()
    
    ws_mk.column_dimensions['A'].width = 5
    ws_mk.column_dimensions['B'].width = 25
    ws_mk.column_dimensions['C'].width = 15
    ws_mk.column_dimensions['D'].width = 8
    ws_mk.column_dimensions['E'].width = 12
    ws_mk.column_dimensions['F'].width = 20
    ws_mk.column_dimensions['G'].width = 18
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="semua_data_{timestamp}.xlsx"'
    wb.save(response)
    
    return response


@login_required(login_url='/admin/login/')
@require_http_methods(["GET"])
def dashboard_stats(request):
    """Return dashboard statistics as JSON for charts"""
    
    # Mahasiswa by Jurusan
    mahasiswa_by_jurusan = list(
        Mahasiswa.objects.values('jurusan')
        .annotate(count=Count('id'))
        .order_by('jurusan')
    )
    
    # Dosen by Homebase
    dosen_by_homebase = list(
        Dosen.objects.values('homebase')
        .annotate(count=Count('id'))
        .order_by('homebase')
    )
    
    # Mata Kuliah by Semester
    matakuliah_by_semester = list(
        MataKuliah.objects.values('semester')
        .annotate(count=Count('id'))
        .order_by('semester')
    )
    
    return JsonResponse({
        'mahasiswa_by_jurusan': mahasiswa_by_jurusan,
        'dosen_by_homebase': dosen_by_homebase,
        'matakuliah_by_semester': matakuliah_by_semester,
    })
    return response